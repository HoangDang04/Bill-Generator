import tkinter as tk
from tkinter import messagebox, Scrollbar
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
import os
import datetime

# Load price table
def get_price_table(sheet_name="Price Table"):
    try:
        wb = load_workbook("bill_data.xlsx")
        sheet = wb[sheet_name]
        price_table = {}
        for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
            price_table[row[0]] = row[1]
        return price_table
    except Exception as e:
        messagebox.showerror("Error", f"Could not load price table: {e}")
        return {}

# Calculate bill total
def calculate_total(order_details):
    price_table = get_price_table()
    total = 0
    details = []

    for quantity, product in order_details:
        price = price_table.get(product, 0)
        total += quantity * price
        details.append((quantity, product, price, quantity * price))
    
    return total, details

# Save the bill to Excel
def save_to_excel(details, total, ship_fee, customer_name, date, comment, sheet_name="Test_bill"):
    if not os.path.exists("bill_data.xlsx"):
        messagebox.showerror("Error", "Excel file not found.")
        return
    
    wb = load_workbook("bill_data.xlsx")
    sheet = wb[sheet_name]
    
    # Parse day and month into a full date in the current year
    day_month = datetime.datetime.strptime(date + f"/{datetime.datetime.now().year}", "%d/%m/%Y")
    formatted_date = day_month.strftime("%a\n%d/%m")  # Format to match example: "Sat\n16/11"

    # Set border and alignment styles
    thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    bold_font = Font(bold=True)
    
    # Check if the date already exists in column 1
    last_date_row = None
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == formatted_date:
            last_date_row = row

    if last_date_row:
        # Find the next empty row below the last row for this date
        new_bill_row = last_date_row + 1
        while sheet.cell(new_bill_row, column=2).value is not None:
            new_bill_row += 1
    else:
        # If date not found, insert a new date row
        if sheet.max_row == 1:
            new_bill_row = sheet.max_row + 1
        else:
            new_bill_row = sheet.max_row + 2
        sheet.insert_rows(new_bill_row)
        sheet.merge_cells(start_row=new_bill_row, start_column=1, end_row=new_bill_row, end_column=7)
        sheet.cell(new_bill_row, 1).value = formatted_date
        sheet.cell(new_bill_row, 1).alignment = Alignment(wrap_text=True)
        sheet.cell(new_bill_row, 1).font = bold_font

        for col in range(1, 8):
            cell = sheet.cell(new_bill_row, col)
            cell.border = thin_border
        sheet.row_dimensions[new_bill_row].height = 30

        new_bill_row += 1  # Next row for bill

    # Write data to the bill
    sheet.cell(new_bill_row, 2).value = customer_name
    sheet.cell(new_bill_row, 3).value = comment

    order_details = "\n".join([f"{q} {p}" for q, p, _, _ in details])
    sheet.cell(new_bill_row, 4).value = order_details
    sheet.cell(new_bill_row, 4).alignment = Alignment(wrap_text=True)


    sheet.cell(new_bill_row, 5).value = total
    sheet.cell(new_bill_row, 6).value = ship_fee
    sheet.cell(new_bill_row, 7).value = total + ship_fee

    # Apply border and center alignment to bill cells
    for col in range(1, 8):
        cell = sheet.cell(new_bill_row, col)
        cell.alignment = center_alignment
        cell.border = thin_border
    
    wb.save("bill_data.xlsx")
    
# GUI Application
def generate_bill():
    order_details = []

    # Updating the total for each item added
    def update_total():
        total, _ = calculate_total(order_details)
        lbl_total.config(text=f"Total: {total}")

    # Adding item in the list
    def add_item():
        try:
            quantity = int(entry_quantity.get())
            product = entry_product.get().strip()
            if not product:
                raise ValueError("Product name cannot be empty.")
            order_details.append((quantity, product))
            listbox_orders.insert(tk.END, f"{quantity} x {product}")
            update_total()
        except ValueError as e:
            messagebox.showerror("Error", str(e))

    # Removing specific item in the list
    def remove_item():
        selected = listbox_orders.curselection()
        if selected:
            index = selected[0]
            listbox_orders.delete(index)
            order_details.pop(index)
            update_total()
        else:
            messagebox.showerror("Error", "No item selected to remove.")

    # Searching item in the list of items in the excel table
    def searching_algo(*args):
        typed_text = entry_product.get().strip().lower()
        listbox_suggestions.delete(0, tk.END)
        if len(typed_text) >= 2: # After input 2 characters to show the items in the searching
            finding = [meal for meal in get_price_table() if typed_text in meal.lower()]
            if finding:
                listbox_suggestions.grid(row = 3, column = 1, padx = 10, pady = 2, sticky = "nsew")
                listbox_suggestions.config(height = len(finding))
                for component in finding:
                    listbox_suggestions.insert(tk.END, component)
            else:
                listbox_suggestions.grid_remove()
        else:
            listbox_suggestions.grid_remove()

    # Selecting item from the searching algorithm
    def select(event):
        if listbox_suggestions.curselection():
            selected_meal = listbox_suggestions.get(listbox_suggestions.curselection())
            entry_product.delete(0, tk.END)
            entry_product.insert(0, selected_meal)
            listbox_suggestions.delete(0, tk.END)

    # Calculating the total of the bill
    def on_calculate():
        try:
        # Fetch and validate shipping fee
            shipping_fee = int(entry_ship.get())
            if shipping_fee < 0:
                raise ValueError("Shipping fee cannot be negative.")

            # Calculate total and details
            total, details = calculate_total(order_details)
            total_with_shipping = total + shipping_fee
        
            # Save the bill
            save_to_excel(details, total, shipping_fee, entry_customer.get(), entry_date.get(), comment.get("1.0", "end-1c"))
        
            # Display popup for final total
            messagebox.showinfo(
                "Bill Generated",
                f"Your total is {total_with_shipping}\n\nBill saved to Excel successfully!"
            )
        
            # Update the total label with shipping fee
            lbl_total.config(text=f"Total: {total_with_shipping}")
        except ValueError as e:
            messagebox.showerror("Error", str(e))
    
    # Open excel file
    def open_excel_file():
        try:
            excel_path = "bill_data.xlsx"
            os.startfile(excel_path)
        except Exception as e:
            messagebox.showerror("Error", "Cannot find excel file")
    
    # Name of the app 
    root = tk.Tk()
    root.title("Billing App")
    
    # Customer name box
    tk.Label(root, text="Customer Name:").grid(row=0, column=0, padx=10, pady=5)
    entry_customer = tk.Entry(root)
    entry_customer.grid(row=0, column=1, padx=10, pady=5)
    
    # Date box
    tk.Label(root, text="Date:").grid(row=1, column=0, padx=10, pady=5)
    entry_date = tk.Entry(root)
    entry_date.grid(row=1, column=1, padx=10, pady=5)
    
    # # Quantity box
    # tk.Label(root, text="Quantity:").grid(row=2, column=0, padx=10, pady=5)
    # entry_quantity = tk.Entry(root)
    # entry_quantity.grid(row=2, column=1, padx=10, pady=5)

    # Meal name box
    tk.Label(root, text="Meal Name:").grid(row=2, column=0, padx=10, pady=5)

    adding_stuff = tk.Frame(root)
    adding_stuff.grid(row=2, column=1, columnspan=3, padx=10, pady=5)

    entry_quantity = tk.Entry(adding_stuff)
    entry_quantity.grid(row=1, column=0, padx=2, pady=5)
    entry_quantity.config(width = int(1/4 * entry_quantity.cget("width")))

    entry_product = tk.Entry(adding_stuff)
    entry_product.grid(row=1, column=1, padx=10, pady=5, sticky="e")

    tk.Button(adding_stuff, text="Add Item", command=add_item).grid(row=1,column=2, padx=10, pady=5, sticky = "w")

    # Items showing from the algorithm
    listbox_suggestions = tk.Listbox(root, height=0)  # Initially hidden
    listbox_suggestions.grid(row=3, column=1, padx=10, pady=2, sticky="nsew")
    listbox_suggestions.grid_remove()  # Hide until needed

    listbox_suggestions.bind("<<ListboxSelect>>", select)
    entry_product.bind("<KeyRelease>", searching_algo)

    # All items you have ordered
    listbox_orders = tk.Listbox(root, height=10, width=40)
    listbox_orders.grid(row=4, column=1, padx=10, pady=5)

    scrollbar = Scrollbar(root, orient="vertical", command=listbox_orders.yview)
    scrollbar.grid(row=4, column=2, sticky="ns")
    listbox_orders.config(yscrollcommand=scrollbar.set)

    # Removing item you dont want
    tk.Button(root, text="Remove Item", command=remove_item).grid(row=5, column=1, pady=5)
    
    # Any other comments for the bills
    tk.Label(root, text="Comment:").grid(row=6, column=0, padx=10, pady=5)
    comment = tk.Text(root, width=30, height=3)
    comment.grid(row=6, column=1, pady=5)
    
    # How much is the shipping
    tk.Label(root, text="Shipping Fee:").grid(row=7, column=0, padx=10, pady=5)
    entry_ship = tk.Entry(root)
    entry_ship.grid(row=7, column=1, padx=10, pady=5)
    
    # Generating the total of the bill
    lbl_total = tk.Label(root, text="Total: 0")
    lbl_total.grid(row=8, column=1, pady=10)
    
    # Calculating the whole order
    tk.Button(root, text="Open Excel Datasheet", command=open_excel_file).grid(row=9, column=0, padx=10, pady=10)
    tk.Button(root, text="Generate Bill", command=on_calculate).grid(row=9, column=1, pady=10)
    
    root.mainloop()

# Run the app
generate_bill()